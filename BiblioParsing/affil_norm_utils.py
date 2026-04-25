"""Module of useful functions for normalization of authors' affiliations.
"""

__all__ = ['build_affils_useful_dicts',
           'build_norm_raw_affils_dict',
           'read_affil_types',
           'read_towns_per_country',
           ]


# Standard library imports
import re
from pathlib import Path

# 3rd party imports
import openpyxl
import pandas as pd

# Local library imports
import BiblioParsing as bp
import BiblioParsing.affiliations_globals as bp_ag
import BiblioParsing.general_globals as bp_gg
import BiblioParsing.regex_globals as bp_rg
from BiblioParsing.parsing_utils import dict_print
from BiblioParsing.parsing_utils import remove_special_symbol
from BiblioParsing.parsing_utils import rationalize_town_names
from BiblioParsing.parsing_utils import set_address_uniform_words


def _build_words_set(raw_aff, verbose=False):
    """Builds sets of words from a raw affiliation after standardization of words and symbols, 
    removing special symbols, adding missing spaces and droping small words.

    Args:
        raw_aff (str): The raw affiliation used to build the sets of words.
        verbose (bool): If true, variables are printed for code control (default: False).
    Returns:
        (tuple): Tuple of to sets of words; The first set is the canonical set of words \
        issuing from the string 'raw_aff'; The second set is an added set if some specific \
        accronyms are present in the first set of words.
    """
    # Setting substitution templates for searching small words or acronyms
    word_to_drop_template = bp_rg.AFFIL_WORD_TO_DROP_TEMPLATE

    # Removing accents and spaces at ends
    raw_aff_mod = remove_special_symbol(raw_aff, only_ascii=False, strip=True)

    # Uniformizing words
    std_raw_aff = set_address_uniform_words(raw_aff_mod)
    std_raw_aff = std_raw_aff.lower()

    # Uniformizing dashes
    std_raw_aff = std_raw_aff.translate(bp_gg.DASHES_CHANGE)

    # Uniformizing apostrophes
    std_raw_aff = std_raw_aff.translate(bp_gg.APOSTROPHE_CHANGE)

    # Uniformizing symbols
    std_raw_aff = std_raw_aff.translate(bp_gg.SYMB_CHANGE)

    # Droping particular symbols
    std_raw_aff = std_raw_aff.translate(bp_gg.SYMB_DROP)
    if verbose:
        print('       std_raw_aff:', std_raw_aff)

    # Building the corresponding set of words to std_raw_aff
    raw_aff_words_set = set(std_raw_aff.strip().split(' '))

    # Managing missing space in raw affiliations related
    # to particuliar institutions cases such as UMR or U followed by digits
    std_raw_aff_add = ""
    for accron in bp_ag.MISSING_SPACE_ACRONYMS:
        re_accron = re.compile(word_to_drop_template.substitute({"word":accron}))
        if re.search(re_accron, std_raw_aff.lower()) and len(raw_aff_words_set)==2:
            std_raw_aff_add = "".join(std_raw_aff.split(" "))

    # Droping small words
    for word_to_drop in bp_ag.SMALL_WORDS_DROP:
        re_drop_words = re.compile(word_to_drop_template.substitute({"word":word_to_drop}))
        if re.search(re_drop_words, std_raw_aff.lower()):
            raw_aff_words_set = raw_aff_words_set - {word_to_drop}

    # Updating raw_aff_words_set_list using std_raw_aff_add
    raw_aff_words_set_add = {}
    if std_raw_aff_add:
        raw_aff_words_set_add = set(std_raw_aff_add.split(' '))
    return raw_aff_words_set, raw_aff_words_set_add


def _build_words_sets_list(raw_aff_list, verbose=False):
    """Builds a list of words sets from a list of raw affiliations.

    Args:
        raw_aff_list (list): The list of raw affiliations as strings.
        verbose (bool): If true, variables are printed for code control (default: False).
    Returns:
        (list): List of words sets.
    """
    raw_aff_words_sets_list = []
    for raw_aff in raw_aff_list:
        if raw_aff and raw_aff!=' ':
            # Building the set of words for raw affiliation
            raw_aff_words_set, raw_aff_words_set_add = _build_words_set(raw_aff, verbose)

            # Updating the list of words sets with the set raw_aff_words_set
            raw_aff_words_sets_list.append(raw_aff_words_set)

            # Updating the list of words sets using the set raw_aff_words_set_add
            if raw_aff_words_set_add:
                raw_aff_words_sets_list.append(raw_aff_words_set_add)

    return raw_aff_words_sets_list


def build_norm_raw_affils_dict(country_affiliations_file_path=None, verbose=False):
    """Builds a dict keyed by country and the value per country is a dict keyed 
    by normalized affiliation and valued by a list of sets of words representing 
    the raw affiliations corresponding to the normalized affiliation.

    Args:
        country_affiliations_file_path (path): Full path to the file of normalized affiliations \
        with they possible corresponding raw affiliation built by the user"; if None, it is set \
        using the 'COUNTRY_AFFILIATIONS_FILE' and 'REP_UTILS' globals.
        verbose (bool): If true, variables are printed for code control (default: False).
    Returns:
        (dict): The built dict.
    """
    # Setting useful column names
    norm_affil_col = bp_ag.AFFIL_COL_NAMES['norm_affil_col']

    # Setting the path for the 'Country_affiliations.xlsx' file
    if not country_affiliations_file_path:
        rep_utils_path = Path(bp.__file__).parent / Path(bp_gg.REP_UTILS)
        country_affils_file = bp_ag.AFFIL_DEFAULT_FILES_DIC['country_affils_file']
        country_affiliations_file_path = rep_utils_path / Path(country_affils_file)

    # Reading the 'Country_affiliations.xlsx' multisheet XLSX file in a dict
    wb = openpyxl.load_workbook(country_affiliations_file_path)
    country_aff_dict = pd.read_excel(country_affiliations_file_path, sheet_name=wb.sheetnames)

    norm_raw_aff_dict = {}
    for country_aff_item in country_aff_dict.items():
        country = country_aff_item[0]
        norm_raw_aff_df = country_aff_item[1]
        norm_raw_aff_nb = len(norm_raw_aff_df[norm_affil_col])

        if verbose:
            print('Country:', country)
            print('Number of normalized affiliations:', norm_raw_aff_nb)
            print('\nList of normalized affiliations:', norm_raw_aff_df[norm_affil_col], "\n")

        norm_raw_aff_dict[country] = {}
        for num, norm_aff in enumerate(norm_raw_aff_df[norm_affil_col]):
            norm_aff = norm_aff.strip()
            raw_aff_list = [item for item in list(norm_raw_aff_df.loc[num])[1:] if not(pd.isnull(item)) is True]

            if verbose:
                print(f"\n\n{str(num)}- Normalized affiliation: {norm_aff}")
                print('   Raw affiliations list:', raw_aff_list, "\n")

            norm_raw_aff_dict[country][norm_aff] = _build_words_sets_list(raw_aff_list, verbose)

            if verbose:
                print(f"   norm_raw_aff_dict[{country}][{norm_aff}]: {norm_raw_aff_dict[country][norm_aff]}\n")

    return norm_raw_aff_dict


def read_affil_types(affil_types_file_path=None):
    """Builds a dict keyed by normalized affiliations types and the value per type 
    is the order level of the type.

    Args:
        affil_types_file_path (path): The full path to the file of ordered affiliations types; \
        if None, it is set using the 'INST_TYPES_FILE' and 'REP_UTILS' globals.
    Returns:
        (dict): The built dict.
    """
    # Setting useful column names
    level_col, short_col = bp_ag.AFFIL_TYPES_USECOLS

    # Setting the full path for the file of ordered institutions types
    if not affil_types_file_path:
        rep_utils_path = Path(bp.__file__).parent / Path(bp_gg.REP_UTILS)
        affil_types_file = bp_ag.AFFIL_DEFAULT_FILES_DIC['affil_types_file']
        affil_types_file_path = rep_utils_path / Path(affil_types_file)

    # Reading the file in a dataframe
    affil_types_df = pd.read_excel(affil_types_file_path, usecols=[level_col, short_col])

    levels = list(affil_types_df[level_col])
    abbreviations = list(affil_types_df[short_col])
    aff_type_dict = dict(zip(abbreviations, levels))

    return aff_type_dict


def read_towns_per_country(country_towns_file=None, country_towns_folder_path=None):
    """Builds dict keyed by countries and valued by a list of towns of the country.

    It uses the functions `rationalize_town_names`and `remove_special_symbol`
    imported from the `BiblioParsing.BiblioParsingUtils` module.

    Args:
        country_towns_file (str): File name of the list of towns per country.
        country_towns_folder_path (path): The full path to the folder \
        of the 'country_towns_file' file.
    Returns:
        (dict): The built dict.
    """
    # Setting the path of the file of towns par country
    if not country_towns_folder_path:
        country_towns_folder_path = Path(bp.__file__).parent / Path(bp_gg.REP_UTILS)
    if not country_towns_file:
        country_towns_file = bp_ag.AFFIL_DEFAULT_FILES_DIC['country_towns_file']
    file_path = country_towns_folder_path / Path(country_towns_file)

    # Reading the file of towns per country in a dict of dataframes
    wb = openpyxl.load_workbook(file_path)
    dfs_dict = pd.read_excel(file_path, sheet_name=wb.sheetnames)

    towns_dict = {x[0]:x[1]['Town name'].tolist() for x in dfs_dict.items()}
    for country in towns_dict.keys():
        list_towns = []
        for town in towns_dict[country]:
            town = town.lower()
            town = rationalize_town_names(town, dic_town_symbols=bp_ag.DIC_TOWN_SYMBOLS,
                                          dic_town_words=bp_ag.DIC_TOWN_WORDS)
            town = remove_special_symbol(town, only_ascii=False, strip=False)
            town = town.strip()
            list_towns.append(town)
        towns_dict[country] = list_towns
    return towns_dict


def _check_norm_raw_affils_dict(affil_types_dict, norm_raw_affils_dict):
    wrong_affil_types_dict = {}
    affil_types_set = set(affil_types_dict.keys())
    for country, country_dict in norm_raw_affils_dict.items():
        norm_affil_types_set = {norm_affil.split(' ')[-1] for norm_affil in country_dict.keys()}
        norm_affil_types_in = affil_types_set.intersection(norm_affil_types_set)
        norm_affil_types_out = norm_affil_types_set - norm_affil_types_in
        if norm_affil_types_out:
            wrong_affil_types_dict[country] = list(norm_affil_types_out)
    return wrong_affil_types_dict


def build_affils_useful_dicts(affil_params_dic):
    """Builds useful data for normalization of authors' affiliations.

    The data are built as 3 dicts through the `read_affil_types`, `build_norm_raw_affils_dict` 
    and `read_towns_per_country` functions of the same module.

    Args:
        affil_params_dic (dict): Keyed by ['affil_types_file_path', 'country_affils_file_path', \
        'country_towns_folder_path', 'country_towns_file'] and valued by the user as the full path to the data \
        per country of raw affiliations per normalized one, the full path to the data of affiliations-types \
        used to normalize the affiliations, the name of the file of the data of towns per country and the full \
        path to the folder where these are available.
    Returns:
        (tup): The 3 built dicts.
    Notes:
        When the 'country_affils_file_path', 'affil_types_file_path', 'country_towns_folder_path' \
        and 'country_towns_file' parameters are set to None, the values are defined by default internally to \
        the `build_norm_raw_affils_dict`, `read_affil_types` and `read_towns_per_country` functions.
    """
    (affil_types_file_path, country_affils_file_path,
     country_towns_folder_path, country_towns_file) = [None] * 4
    if affil_params_dic:
        params_keys = ['affil_types_file_path', 'country_affils_file_path',
                       'country_towns_folder_path', 'country_towns_file']
        (affil_types_file_path, country_affils_file_path,
         country_towns_folder_path, country_towns_file) = [affil_params_dic[key] for key in params_keys]

    # Building the useful data for affiliations normalization
    affil_types_dict = read_affil_types(affil_types_file_path=affil_types_file_path)
    norm_raw_affils_dict = build_norm_raw_affils_dict(country_affiliations_file_path=country_affils_file_path,
                                                     verbose=False)
    towns_dict = read_towns_per_country(country_towns_file=country_towns_file,
                                        country_towns_folder_path=country_towns_folder_path)

    # Checking affiliations-types in 'norm_raw_affils_dict' dict
    wrong_affil_types_dict = _check_norm_raw_affils_dict(affil_types_dict, norm_raw_affils_dict)
    if wrong_affil_types_dict:
        print("\nWARNING: Uncorrect normalized-affiliation types found in the file: "
              f"\n         {country_affils_file_path}"
              "\n\n         Please, correct the following affiliation types:")
        dict_print(wrong_affil_types_dict)

    # Building returned dict
    affil_dicts = {'affil_types_dict'      : affil_types_dict,
                   'norm_raw_affils_dict'  : norm_raw_affils_dict,
                   'towns_dict'            : towns_dict,
                   'wrong_affil_types_dict': wrong_affil_types_dict,
                  }
    return affil_dicts
