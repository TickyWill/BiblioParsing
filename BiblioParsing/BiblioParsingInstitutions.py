__all__ = ['address_inst_full_list',
           'build_norm_raw_affiliations_dict',
           'build_norm_raw_institutions',
           'extend_author_institutions',
           'read_inst_types',
           'read_towns_per_country',
           ]


# Standard library imports
import re
from pathlib import Path
from collections import namedtuple
from string import Template

# 3rd party imports
import openpyxl
import pandas as pd

# Local library imports
import BiblioParsing as bp
import BiblioParsing.BiblioGeneralGlobals as bp_gg
import BiblioParsing.BiblioSpecificGlobals as bp_sg
from BiblioParsing.BiblioParsingUtils import remove_special_symbol
from BiblioParsing.BiblioParsingUtils import rationalize_town_names
from BiblioParsing.BiblioParsingUtils import build_item_df_from_tup
from BiblioParsing.BiblioParsingUtils import standardize_address


def _set_norm_affiliations_cols():
    """Builds 2 dict setting columns lists and selected columns names 
    for the process of parsing author affiliations and getting their 
    normalized affiliations.

    Returns:
        (tup): (A dict valued by column-names lists defined by the 'COL_NAMES' global, \
        A dict valued by column names of parsing results defined by the 'COL_NAMES' global).
    """
    
    cols_lists_dic = {'country_cols_list'  : bp_sg.COL_NAMES['country'],
                      'inst_cols_list'     : bp_sg.COL_NAMES['institution'],
                      'auth_inst_cols_list': bp_sg.COL_NAMES['auth_inst'],
                     }
    
    cols_dic = {'pub_id_col'          : bp_sg.COL_NAMES['pub_id'],
                'address_col'         : bp_sg.COL_NAMES['address'][2],
                'country_col'         : bp_sg.COL_NAMES['country'][2],
                'institution_col'     : bp_sg.COL_NAMES['institution'][2],
                'norm_institution_col': bp_sg.COL_NAMES['auth_inst'][4],
               }
    return cols_lists_dic, cols_dic


def _get_norm_affiliations_list(country, affiliations_list, norm_raw_aff_dict, 
                                aff_type_dict, verbose=False):
    """ToDo: docstring fill.
    """
    # Setting useful regex template
    # Capturing for instence "word" in "word of set"
    # or " word" in "set with word"
    # or "word" in "Azert Word Azerty"
    set_words_template = Template(r'[\s]$word[\s)]' + '|'
                                  + r'[\s]$word$$' + '|'
                                  + r'^$word\b')

    address_norm_affiliations_list = []
    address_unknown_affiliations_list = [] 
    for affiliation in affiliations_list:
        if verbose: 
            print(' -', affiliation)
        norm_affiliation_list = []

        # Removing accents and converting to lower case
        aff_mod = remove_special_symbol(affiliation, only_ascii=False, strip=True)
        aff_mod = aff_mod.lower()
        if verbose:
            print()
            print('aff_mod:', aff_mod)
            print()

        # Searching for words set in affiliation
        for num, norm_aff in enumerate(norm_raw_aff_dict[country].keys()):

            if verbose:
                print()
                print(str(num) + ' norm_aff:', norm_aff)
                print()

            for words_set in norm_raw_aff_dict[country][norm_aff]:
                if verbose :print('  words_set:', words_set)
                words_set_tags = []
                for word in words_set:
                    if verbose: print('    word:', word)
                    re_search_words = re.compile(set_words_template.substitute({"word":word}))
                    if re.search(re_search_words,aff_mod) :
                        words_set_tags.append('true')
                        if verbose: print('     words_set_tags:', words_set_tags)
                    else:
                        words_set_tags.append('false')
                        if verbose: print('     words_set_tags:', words_set_tags)

                if 'false' not in words_set_tags:               
                    norm_affiliation_list.append(norm_aff)

                if verbose: 
                    print('  words_set_tags:',words_set_tags)
                    print('  norm_affiliation_list:', norm_affiliation_list)
                    print()

        if verbose: print('  norm_affiliation_list:', norm_affiliation_list)

        if norm_affiliation_list==[]:
            address_unknown_affiliations_list.append(affiliation)

        address_norm_affiliations_list = address_norm_affiliations_list + norm_affiliation_list 

    address_norm_affiliations_set = set(address_norm_affiliations_list)
    if verbose: 
        print('address_norm_affiliations_list:', address_norm_affiliations_list)
        print('address_norm_affiliations_set:     ', address_norm_affiliations_set)

    paris_nb = 0
    for norm_aff in address_norm_affiliations_set:
        if 'Univ' in norm_aff and 'Paris' in norm_aff:
            paris_nb += 1

    if paris_nb>1 and 'Paris-Cité Univ' in address_norm_affiliations_set:
        address_norm_affiliations_set = address_norm_affiliations_set - {'Paris-Cité Univ'}
    if verbose: 
        print('address_norm_affiliations_set:     ',address_norm_affiliations_set)

    idx_dict = dict(zip(aff_type_dict.keys(), [0 ]* len(aff_type_dict.keys())))
    norm_aff_pos_list = []
    address_norm_affiliation_dict = {}      
    for norm_aff in address_norm_affiliations_set:        
        norm_aff_type = norm_aff.split(' ')[-1]
        if verbose:
            print('norm_aff_type:', norm_aff_type)
            print('str(idx_dict[norm_aff_type]):', str(idx_dict[norm_aff_type]))

        norm_aff_pos = str(aff_type_dict[norm_aff_type]) + str(idx_dict[norm_aff_type])
        if verbose: 
            print('norm_aff_pos init:',norm_aff_pos)
            print('norm_aff_pos_list init:', norm_aff_pos_list)
        if int(norm_aff_pos) in norm_aff_pos_list: 
            idx_dict[norm_aff_type] += 1

        norm_aff_pos = str(aff_type_dict[norm_aff_type]) + str(idx_dict[norm_aff_type])
        if verbose: 
            print('norm_aff_pos end:',norm_aff_pos)
            print('idx_dict[norm_aff_type]:', idx_dict[norm_aff_type])

        norm_aff_pos_list.append(int(norm_aff_pos))
        if verbose: 
            print('norm_aff_pos_list end:', norm_aff_pos_list)
            print()

        address_norm_affiliation_dict[norm_aff_pos] = norm_aff 

    if verbose:
        print('address_norm_affiliation_dict:     ', address_norm_affiliation_dict) 
    norm_aff_pos_list.sort()

    address_norm_affiliation_list = [None] * len(address_norm_affiliations_set)
    for idx in range(len(norm_aff_pos_list)):
        address_norm_affiliation_list[idx] = address_norm_affiliation_dict[str(norm_aff_pos_list[idx])]

    return (address_norm_affiliation_list, address_unknown_affiliations_list)


def _build_address_affiliations_lists(std_address, norm_raw_aff_dict, aff_type_dict, 
                                      towns_dict, drop_status, verbose=False):
    """Builds the list of normalized affiliations for a standardized address.

    It also returns the country and the unknown affiliations for this address. 
    To do that, it uses the `_get_affiliations_list` and `_get_norm_affiliations_list` 
    internal functions.
    
    Args:
        std_address (str): The standardized address for which the list of normalized affiliations is built.
        norm_raw_aff_dict (dict): A dict used for the normalization of the institutions names, \
        with the normalized names as keys and the raw names as values and built by \
        the `build_norm_raw_affiliations_dict` function of the same module.
        aff_type_dict (dict): A dict used to validate the normalized names of the institutions and \
        to set the order of these names by institution type; it is built by the function `read_inst_types` \
        function of the same module.
        towns_dict (dict): A dict used to identifie the towns in the address in order to drop them; \
        it is built by the `read_towns_per_country` function of the same module.
        drop_status (bool): If true, droping items are searched to drop chunks from the address.
        verbose (bool): True for allowing control prints (default: False).
    Returns:
        (tuple): A tuple of 3 items; first item is the country as string; \
        second item is the list of normalized affiliations; \
        third item is the list of unknown affiliations.
    """
    if verbose:
        print()
        print('Standardized address:              ', std_address)

    return_tup = _get_affiliations_list(std_address, towns_dict, drop_status=drop_status, verbose=False)
    country, affiliations_list, affiliations_drop = return_tup
    affiliations_list_mod = [affiliation.translate(bp_gg.SYMB_CHANGE) for affiliation in affiliations_list]
    
    if verbose:
        print()
        print('Country:                           ', country)
        print()
        print('Affiliations list:                 ', affiliations_list)
        print('Modified affiliations list:        ', affiliations_list_mod)
        print('Affiliations dropped:              ', affiliations_drop) 

    if country in norm_raw_aff_dict.keys():
        return_tup = _get_norm_affiliations_list(country, affiliations_list_mod, norm_raw_aff_dict, 
                                                 aff_type_dict, verbose=False)
        address_norm_affiliation_list, address_unknown_affiliations_list = return_tup
    else:
        address_norm_affiliation_list = []
        address_unknown_affiliations_list = affiliations_list
    
    return country, address_norm_affiliation_list, address_unknown_affiliations_list


def address_inst_full_list(full_address, norm_raw_aff_dict, aff_type_dict, towns_dict, drop_status):
    """Builds the affiliations list of a full address using the `_build_address_affiliations_lists` 
    internal function of the same module.
    
    Args:
        full_address (str): the full address to be parsed in institutions and country.
        norm_raw_aff_dict (dict): A dict used for the normalization of the institutions names, \
        with the normalized names as keys and the raw names as values and built by \
        the `build_norm_raw_affiliations_dict` function of the same module.
        aff_type_dict (dict): A dict used to validate the normalized names of the institutions and \
        to set the order of these names by institution type; it is built by the function `read_inst_types` \
        function of the same module.
        towns_dict (dict): A dict used to identifie the towns in the address in order to drop them; \
        it is built by the `read_towns_per_country` function of the same module. 
        drop_status (bool): If true, droping items are searched to drop chunks from the address.
    Returns:
        (namedtuple): A tuple of two strings; the first is the joined list of normalized institutions \
        names found in the full address; the second is the joined list of raw institutions names \
        of the full address with no fully corresponding normalized names.
    """
    inst_full_list_ntup = namedtuple('inst_full_list_ntup', ['norm_inst_list','raw_inst_list'])

    aff_list_tup = _build_address_affiliations_lists(full_address, norm_raw_aff_dict,
                                                     aff_type_dict, towns_dict, drop_status,
                                                     verbose = False)
    country, norm_inst_full_list, raw_inst_full_list = aff_list_tup

    if raw_inst_full_list:
        raw_inst_full_list_str = ";".join(raw_inst_full_list)       
    else:
        raw_inst_full_list_str = bp_sg.EMPTY 

    # Building a string from the final list of normalized institutions without duplicates
    norm_inst_full_list = list(set(norm_inst_full_list))
    if norm_inst_full_list:
        norm_inst_full_list_str = ";".join(norm_inst_full_list)
    else:
        norm_inst_full_list_str = bp_sg.EMPTY 

    # Setting the namedtuple to return
    inst_full_list_tup =  inst_full_list_ntup(norm_inst_full_list_str, raw_inst_full_list_str)

    return inst_full_list_tup


def extend_author_institutions(item_df, inst_filter_list):
    """Extends the data of authors with affiliation institutions initialy obtained 
    by the parsing of the corpus, with complementary information about institutions
    selected by the user.

    The selection is given by the user through a list of 2-items tuples composed 
    of a normalized institution and the corresponding collumn name. For each normalized 
    institution, the corresponding column is filled with 1 for each of the author 
    affiliated to this institution. Otherwise, it is filled with 0.

    Args:
        item_df (dataframe): The data of authors with affiliation institutions.
        inst_filter_list (list): The list of tuples selected by the user.
    Retruns:
        (dataframe): The extended data with the columns given by the user.
    """
    # Internal function
    def _build_complement_inst_list(inst_names_list, institutions):
        complement_inst_list = []
        for inst in inst_names_list:
            if inst in institutions:
                complement_inst_list.append(1)
            else:
                complement_inst_list.append(0)
        return complement_inst_list

    
    # Setting useful column names
    cols_lists_dic, cols_dic = _set_norm_affiliations_cols()
    read_usecols = cols_lists_dic['auth_inst_cols_list'][0:5]
    norm_institution_col = cols_dic['norm_institution_col']
    temp_col = "temp_col"

    # Getting the useful columns of the item df
    item_df = item_df[read_usecols]

    # Setting an institution name for each of the institutions indicated in the institutions filter
    inst_names_list = [f'{x[0]}' for x in inst_filter_list]
    inst_col_list = [f'{x[1]}' for x in inst_filter_list]

    # Building a list of 0 or 1 in 'temp_col' column added to the initial data using "inst_filter_list"
    item_dg = item_df.copy()
    item_dg[temp_col] = item_dg.apply(lambda row: _build_complement_inst_list(inst_names_list,
                                                                              row[norm_institution_col]),
                                      axis=1)
    item_dg.reset_index(inplace=True, drop=True)

    # Distributing the value lists of 'temp_col' column in a dataframe
    # into columns which names are in 'inst_col_list' list
    inst_split_df = pd.DataFrame(item_dg[temp_col].sort_index().to_list(),
                                 columns=inst_col_list)

    # Extending the initial data with the previously built data from 'temp_col' column
    new_item_df = pd.concat([item_dg, inst_split_df], axis=1)

    # Droping the temp_col column which is no more useful
    new_item_df.drop([temp_col], axis=1, inplace=True)
    return new_item_df


def  _search_droping_bp(params_list, verbose=False):
    """Searches in the passed stringfor words begenning with 'bp' followed 
    by digits using a non case sensitive regex.

    Args:
        params_list (list): Composed of the string where the words are searched \
        after being converted to lower case, of the string (unused) \
        that contains the country and of the dict used to identify the towns \
        in the address (unused).
        verbose (bool): True for allowing control prints (default: False).
    Returns:
        (list): Composed of one boolean; True if a word begenning with 'bp' \
        followed by digits is found.
    """
    # Setting useful params values from params_list
    text = params_list[0]

    # Setting the regex for capturing, for instence "bp12" in "azert BP12 yui_OP"
    # capturing " bp 156X" in " bp 156X azert" or capturing "08bp" in "azert 08BP yui_OP".
    re_bp = re.compile(r'\bbp\s?\d+[a-z]?\b' + '|' + r'\b\d+bp\b')

    flag = False
    result = re.search(re_bp, text.lower())
    if result is not None:
        if verbose:
            print('Droping word is postal-box abbreviation')
        flag = True
    return [flag]


def _search_droping_digits(params_list, verbose=False):
    """Searches in the passed string for words similar to zip codes except those 
    begenning with a prefix from the global 'KEEPING_PREFIX' followed by 3 or 4 digits 
    using case-sensitive regexes. 

    The regex for zip-codes search uses the global 'ZIP_CODES' dict for countries 
    from 'ZIP_CODES.keys()'. 
    Specific regex are set for 'United Kingdom', 'Canada' and 'United States'. 

    Args:
        params_list (list): Composed of the string where the words are searched \
        after being converted to lower case, of the string that contains the country \
        and of the dict used to identify the towns in the address (unused).
        verbose (bool): True for allowing control prints (default: False).
    Returns:
        (list): Composed of one boolean; True if a word different from those \
        begenning with a prefix from the global 'KEEPING_PREFIX' followed \
        by 3 or 4 digits is found.
    """
    # Setting useful params values from params_list
    text, country, _ = params_list
    
    # Setting regex for zip-codes search
    pattern = ''
    if country=='United Kingdom':
        # Capturing: for instence, " BT7 1NN" or " WC1E 6BT" or " G128QQ"
        #            " a# #a", " a# #az", " a# ##a", " a# ##az",
        #            " a##a", " a##az", " a###a", " a###az",
        #
        #            " a#a #a", " a#a #az", " a#a ##a", " a#a ##az",
        #            " a#a#a", " a#a#az", " a#a##a", " a#a##az",
        #
        #            " a## #a", " a## #az", " a## ##a", " a## ##az",
        #            " a###a", " a###az", " a####a", " a####az",
        #            
        #            " a##a #a", " a##a #az", " a##a ##a", " a##a ##az",
        #            " a##a#a", " a##a#az", " a##a##a", " a##a##az",
        #            
        #            " az# #a", " az# #az", " az# ##a", " az# ##az",
        #            " az##a", " az##az", " az###a", " az###az",
        #
        #            " az#a #a", " az#a #az", " az#a ##a", " az#a ##az",
        #            " az#a#a", " az#a#az", " az#a##a", " az#a##az",
        #
        #            " az## #a", " az## #az", " az## ##a", " az## ##az",
        #            " az###a", " az###az", " az###a", " az####az",
        #
        #            " az##a #a", " az##a #az", " az##a ##a", " az##a ##az",
        #            " az##a#a", " az##a#az", " az##a#a", " az##a##az",
        pattern = r'^\s?[a-z]{1,2}\d{1,2}[a-z]{0,1}\s?\d{1,2}[a-z]{1,2}$'

    elif country=='United States' or country=='Canada':
        # Capturing: for instence, " NY" or ' NI BT48 0SG' or " ON K1N 6N5" 
        #            " az" or " az " + 6 or 7 characters in 2 parts separated by spaces
        pattern = r'^\s?[a-z]{2}$' + '|' + r'^\s?[a-z]{2}\s[a-z0-9]{3,4}\s[a-z0-9]{2,3}$'

    elif country in bp_gg.ZIP_CODES.keys():
        letters_list = bp_gg.ZIP_CODES[country]['letters']
        digits_list = bp_gg.ZIP_CODES[country]['digits']

        if letters_list or digits_list:
            zip_template = Template(r'\b($zip_letters)[\s-]?(\d{$zip_digits})\b')

            letters_join = '|'.join(letters_list) if len(letters_list) else ''
            pattern_zip_list = [zip_template.substitute({"zip_letters": letters_join,
                                                         "zip_digits":digits})
                                for digits in digits_list]
            pattern = '|'.join(pattern_zip_list)        
    else:
        print('country not found:', country)
        flag = False
        return [flag]

    zip_result = False
    if pattern: 
        re_zip = re.compile(pattern)
        if re.search(re_zip,text.lower()): zip_result = True   

    # Setting search regex of embedding digits
    # In first part, for captuting, for instence, " 1234" in "azert 1234-yui_OP"
    # or " 1" in "azert 1-yui_OP" or " 1-23" in "azert 1-23-yui"
    # Or, in second part, capturing, for instence, "azert12" in "azert12 UI_OPq"
    # or "azerty1234567" in "azerty1234567 ui_OPq"
    re_digits = re.compile(r'\s?\d+(-\d+)?\b' + '|' + r'\b[a-z]+(-)?\d{2,}\b')

    # Setting search regex of keeping-prefix
    # for instence, capturing "umr1234" in "azert UMR1234 YUI_OP"
    # or "fr1234" in "azert-fr1234 Yui_OP".
    prefix_template = Template(r'\b$prefix[-]?\d{4}\b')
    pattern_prefix_list = [prefix_template.substitute({"prefix": prefix})
                           for prefix in bp_sg.KEEPING_PREFIX]   
    re_prefix = re.compile('|'.join(pattern_prefix_list))             


    prefix_result = False if (re.search(re_prefix,text.lower()) is None) else True
    if prefix_result and verbose: print('Keeping prefix: True')

    digits_result = False if (re.search(re_digits,text.lower()) is None) else True

    flag = False
    if not prefix_result and (zip_result or digits_result):    
        if verbose:
            print('Droping word is a zip code') if zip_result else print('Droping word is a digits code')   
        flag = True

    return [flag]


def _search_droping_suffix(params_list, verbose=False):
    """Searches in the passed string for words ending by a suffix among 
    those given by the global 'DROPING_SUFFIX' using a templated regex.

    Args:
        params_list (list): Composed of the string where the words are searched \
        after being converted to lower case, of the string (unused) that contains \
        the country and of the dict used to identify the towns in the address (unused).
        verbose (bool): True for allowing control prints (default: False).
    Returns:
        (list): Composed of one boolean; True if a suffix given by the 'DROPING_SUFFIX' \
        global is found.               
    """
    # Setting useful params values from params_list
    text = params_list[0]

    # Setting regex for droping-suffix search
    # For instence, capturing "platz" in "Azertyplatz uiops12"
    # Or, for instence, capturing "-gu" in "Yeongtong-gu"
    droping_suffix_template = Template(r'\B$word\b' + '|' + r'\b$word\b')

    flag = False
    for word_to_drop in bp_sg.DROPING_SUFFIX:
        re_drop_words = re.compile(droping_suffix_template.substitute({"word":word_to_drop}))
        result = re.search(re_drop_words, text.lower())
        if result is not None:
            flag = True
            if verbose:
                print('Droping word contains the suffix:', word_to_drop)
    return [flag]


def _search_droping_town(params_list, verbose=False):
    """Searches in the passed string for words in lower case
    that are towns for each country as given in the passed dict 
    of towns per country.

    Args:
        params_list (list): Composed of the string where the words are searched \
        after being converted to lower case, of the string that contains the country \
        and of the dict used to identify the towns in the address.
        verbose (bool): True for allowing control prints (default: False).
    Returns:
        (list): Composed of one boolean; True if a word listed in the values \
        of the dict 'towns_dict' is equal to the passed string after spaces removal \
        at ends.
    """
    # Setting useful params values from params_list
    text, country, towns_dict = params_list

    flag = False
    text_mod = rationalize_town_names(text.lower())
    if country in towns_dict.keys():
        for word_to_drop in towns_dict[country]:
            if word_to_drop==text_mod.strip(): 
                if verbose:
                    print('Droping word is a town of ', country)
                flag = True
    return [flag]


def _search_droping_words(params_list, verbose=False):
    """Searches in the passed string for isolated words given by the 'FR_DROPING_WORDS' 
    and 'DROPING_WORDS' globals using a templated regex.

    If country is 'France' only the 'FR_DROPING_WORDS' global is used.

    Args:
        params_list (list): Composed of the string where the words are searched \
        after being converted to lower case, of the string that contains the country \
        and of the dict used to identify the towns in the address (unused).
        verbose (bool): True for allowing control prints (default: False).
    Returns:
        (list): Composed of one boolean; True if a word given by the 'DROPING_WORDS' \
        or 'FR_DROPING_WORDS' globals is found.
    """
    # Setting useful params values from params_list
    text, country, _ = params_list

    # Setting templated regex for droping-words search
    # For instence, capturing "avenue" in "12 Avenue Azerty" or " cedex" in "azert cedex"
    # in "12 Avenue Azerty" or " cedex" in "azert cedex"
    droping_words_template = Template(r'[\s(]$word[\s)]' + '|' + r'[\s]$word$$' + '|' + r'^$word\b')

    flag = False
    if country.lower()=='france':
        droping_words_to_search = bp_sg.FR_DROPING_WORDS
    else:
        droping_words_to_search = bp_sg.FR_DROPING_WORDS + bp_sg.DROPING_WORDS

    for word_to_drop in droping_words_to_search:
        re_drop_words = re.compile(droping_words_template.substitute({"word":word_to_drop}))
        result = re.search(re_drop_words, text.lower())
        if result is not None:
            flag = True
            if verbose:
                print('Droping word is the full word:', word_to_drop)
    return [flag]


def _search_keeping_prefix(params_list, verbose=False):                               
    """'Searches in the passed string for prefixes given by the global 'KEEPING_PREFIX' 
    using a templated regex if country is France.

    Args:
        params_list (list): Composed of the string where the words are searched \
        after being converted to lower case, of the string that contains the country \
        and of the dict used to identify the towns in the address (unused).
        verbose (bool): True for allowing control prints (default: False).
    Returns:
        (list): Composed of one boolean; True if a prefix given by the 'KEEPING_PREFIX' \
        global is found.
    """ 
    # Setting useful params values from params_list
    text, country, _ = params_list

    # Setting templated regex for keeping prefixes search            
    keeping_prefix_template = Template(r'\b$prefix\d{3,4}\b') 

    flag = False
    if country.lower()=='france':
        for prefix_to_keep in bp_sg.KEEPING_PREFIX:
            re_keep_prefix = re.compile(keeping_prefix_template.substitute({"prefix":prefix_to_keep}))
            result = re.search(re_keep_prefix, text.lower())
            if result is not None:
                if verbose:
                    print('Keeping word is the prefix:', prefix_to_keep)
                flag = True
    return [flag]


def _search_keeping_words(params_list, verbose=False):
    """Searches in the passed string for isolated words given by the 'KEEPING_WORDS' 
    global using a templated regex.

    Args:
        params_list (list): Composed of the string where the words are searched \
        after being converted to lower case, of the string (unused) that contains \
        the country and of the dict used to identify the towns in the address (unused).
        verbose (bool): True for allowing control prints (default: False).
    Returns:
        (list): Composed of 3 booleans all False if no word given by the 'KEEPING_WORDS' \
        global is found; the first is True if a word given by the 'GEN_KEEPING_WORDS' \
        global is found; the second is True if a word given by the 'BASIC_KEEPING_WORDS' \
        global is found; the third is True if a word given by the 'USER_KEEPING_WORDS' \
        global is found.
    """
    # Setting useful params values from params_list
    text = params_list[0]

    # Setting templated regex for keeoing-words search
    keeping_words_template = Template(r'\b$word\b')

    gen_flag, basic_flag, user_flag = False, False, False
    for word_to_keep in bp_sg.KEEPING_WORDS:
        re_keep_words = re.compile(keeping_words_template.substitute({"word":word_to_keep}))
        result = re.search(re_keep_words, text.lower())
        if result is not None:
            if verbose:
                print('Keeping word is the full word:', word_to_keep)
            if word_to_keep in bp_sg.GEN_KEEPING_WORDS:
                gen_flag = True
            if word_to_keep in bp_sg.BASIC_KEEPING_WORDS:
                basic_flag = True
            if word_to_keep in bp_sg.USER_KEEPING_WORDS:
                user_flag = True
    return [gen_flag, basic_flag, user_flag]


def _search_items(affiliation, country, towns_dict, verbose=False):
    """Searches for several item types in the passed chunck of address after accents removal 
    and converting in lower case even if the search is case sensitive.

    It uses the following internal functions:
        - The function `_search_droping_bp` searches for words that are postal-box numbers such as 'BP54'.
        - The function `_search_droping_digits` searches for words that contains digits such as zip codes \
        which templates are given per country by the 'ZIP_CODES' dict global.
        - The function `_search_droping_suffix` searches for words ending by a suffix among \
        those given by the'DROPING_SUFFIX'  global such as 'strasse' in 'helmholtzstrasse'.
        - The function `_search_droping_town` searches for words that are towns listed \ 
        in the 'towns_dict' dict.
        - The function `_search_droping_words` searches for words given by the 'DROPING_WORDS' global \
        such as 'Avenue'.
        - The function `_search_keeping_words` searches for isolated words given by the 'KEEPING_WORDS' \
        global using a templated regex.
        - The function `_search_keeping_prefix` searches for prefixes given by the 'KEEPING_PREFIX' \
        global using a templated regex.
    
    As a reminder, in a regex:
        - '\b' captures the transition between a non-alphanumerical symbol and an alphanumerical symbol \
        and vice-versa.
        - '\B' captures the transition between two alphanumerical symbols.

    Args:
        affiliation (str): A chunck of a standardized address where droping items are searched.
        country (str): The string that contains the country.
        towns_dict (dict): A dict used to identifie the towns in the address in order to drop them; \
        it is built by the `read_towns_per_country` function of the same module. 
        verbose (bool): True for allowing control prints (default: False).
    Returns:
        (namedtuple): A namedtuple which values are booleans returned by the internal functions \
        that returns a list of booleans that are True if the corresponding searched item is found.
    """
    funct_list = [
                  _search_droping_bp,
                  _search_droping_digits,
                  _search_droping_suffix,
                  _search_droping_town,
                  _search_droping_words,
                  _search_keeping_prefix,
                  _search_keeping_words,
                  ]
    
    found_item_tup = namedtuple('found_item_tup', ['droping_bp',
                                                   'droping_digits',
                                                   'droping_suffix',
                                                   'droping_town',
                                                   'droping_words',
                                                   'keeping_prefix',
                                                   'gen_keeping_words',
                                                   'basic_keeping_words',
                                                   'user_keeping_words',
                                                   ])
    
    affiliation_mod = remove_special_symbol(affiliation, only_ascii=False, strip=False)
    params_list = [affiliation_mod, country, towns_dict]
    flag_list = [funct(params_list, verbose) for funct in funct_list]
    
    # Flattening flag_list
    flag_list = sum(flag_list, [])
    found_item_flags = found_item_tup(*flag_list)
    return found_item_flags


def _get_affiliations_list(std_address, towns_dict, drop_status=True, verbose=False):
    """Extracts first, the country and then, the list of institutions from the standardized 
    address. 
    
    It splits the address in list of chuncks separated by coma or isolated hyphen-minus. 
    The country is present as the last chunk of the spliting. 
    The other chunks are kept as institutions if they contain at least one word among 
    those listed in the global 'KEEPING_WORDS' or if they do not contain any item 
    searched by the function `search_droping_items`. 
    The first chunck is always kept in the final institutions list. 
    The spaces at the ends of the items of the final institutions list are removed.
    
    Args:
        std_address (str): The full address to be parsed in list of institutions and country.
        towns_dict (dict): A dict used to identifie the towns in the address in order to drop them; \
        it is built by the `read_towns_per_country` function of the same module.
        drop_status (bool): If True (default: True), droping items are searched to drop chunks \
        from the address. 
        verbose (bool): True for allowing control prints (default: False).
    Returns:
        (tuple): A tuple composed of 3 items (list of kept chuncks, country and list of dropped chuncks).
    """
    
    # Splitting by coma the standard address in chuncks listed in an initial-affiliations list
    init_raw_affiliations_list = std_address.split(',')

    # Removing the first occurence of chunck duplicates from the initial-affiliations list
    # and putting them in a deduplicated-affiliations list
    drop_aff_idx_list = []
    for idx1, aff1 in enumerate(init_raw_affiliations_list): 
        drop_aff_idx_list.extend([min(idx1, idx2) for idx2, aff2 in enumerate(init_raw_affiliations_list) 
                                  if idx1!=idx2 and aff1==aff2])
    dedup_raw_affiliations_list = []    
    dedup_raw_affiliations_list.extend([aff for idx, aff in enumerate(init_raw_affiliations_list) 
                                        if idx not in set(drop_aff_idx_list)])             

    # Setting country index in raw-affiliations list
    country_pos = -1
    country = dedup_raw_affiliations_list[country_pos].strip()

    # Splitting by special characters the deduplicated chunks and putting them in a raw-affiliations list
    raw_affiliations_list = sum([x.split(' - ') for x in dedup_raw_affiliations_list], [])
    raw_affiliations_list = sum([x.split(' | ') for x in raw_affiliations_list], [])

    if drop_status:
        # Initializing the affiliations list by keeping systematically the first chunck of the full address
        affiliations_list = [raw_affiliations_list[0]]

        # Check affiliations only if length > 3 to avoid keeping affiliations of less than 3 characters
        check_affiliations_list = [aff for aff in raw_affiliations_list[1:] if len(aff)>3]

        if verbose: 
            print('Full standard address:',std_address)
            print('init_raw_affiliations_list:',init_raw_affiliations_list)
            print('dedup_raw_affiliations_list:',dedup_raw_affiliations_list)    
            print('country:', country)
            print('raw_affiliations_list flattened:',raw_affiliations_list)
            print('First affiliation:',dedup_raw_affiliations_list[0])
            print('check_affiliations_list:',check_affiliations_list, "\n")

        # Initializing the list of chuncks to drop from the raw-affiliations list
        affiliations_drop = []                                                                          

        # Searching for chuncks to keep and chuncks to drop in the raw-affiliations list, the first chunck and the country excepted
        if len(check_affiliations_list): 
            if verbose:
                print('Search results\n')
            for index, affiliation in enumerate(check_affiliations_list[:country_pos]):        
                if verbose: 
                    print()
                    print('index:', index, '  affiliation:', affiliation)
                found_item_flags = _search_items(affiliation, country, towns_dict, verbose=verbose)
                if verbose:
                    print('found_item_flags:', found_item_flags)
                add_affiliation_flag = False
                break_id = None
                droping_word_flags = [found_item_flags.droping_bp,
                                      found_item_flags.droping_digits,
                                      found_item_flags.droping_suffix,
                                      found_item_flags.droping_town,
                                      found_item_flags.droping_words]

                keeping_words_flags = [found_item_flags.gen_keeping_words,
                                      found_item_flags.basic_keeping_words,
                                      found_item_flags.user_keeping_words]

                if not any(droping_word_flags):
                    affiliations_list.append(affiliation)
                    add_affiliation_flag = True
                    if verbose:
                        print('No droping item found in:', affiliation, '\n')
                else:                
                    if found_item_flags.droping_bp:
                        affiliations_drop.append(('droping_bp', check_affiliations_list[index:country_pos]))
                        break_id = 'droping_bp'
                        if verbose: 
                            print('Break identification:', break_id, '\n')
                        break
                    if found_item_flags.droping_digits:
                        if country.lower() in ['france', 'algeria']:
                            if not found_item_flags.keeping_prefix and not any(keeping_words_flags):
                                affiliations_drop.append(('droping_digits', check_affiliations_list[index:country_pos]))
                                break_id = 'droping_digits'
                                if verbose:
                                    print('Break identification:', break_id, '\n')
                                break
                            elif found_item_flags.gen_keeping_words:
                                if not add_affiliation_flag: 
                                    affiliations_list.append(affiliation)
                                    add_affiliation_flag = True
                                break_id = 'droping_digits aborted by gen_keeping_words'     
                                if verbose:
                                    print('Break identification:', break_id, '\n')
                            else: 
                                if not add_affiliation_flag: 
                                    affiliations_list.append(affiliation)
                                    add_affiliation_flag = True
                                if found_item_flags.basic_keeping_words: break_id = 'droping_digits aborted by basic_keeping_words'
                                if found_item_flags.user_keeping_words : break_id = 'droping_digits aborted by user_keeping_words'
                                if found_item_flags.keeping_prefix: break_id = 'droping_digits aborted by keeping_prefix'
                                if verbose:
                                    print('Break identification:', break_id, '\n')                    
                        else:
                            if not found_item_flags.gen_keeping_words and not found_item_flags.user_keeping_words:
                                affiliations_drop.append(('droping_digits',check_affiliations_list[index:country_pos]))
                                break_id = 'droping_digits'
                                if verbose:
                                    print('Break identification:', break_id, '\n')
                                break
                            elif found_item_flags.droping_words:
                                affiliations_drop.append(('droping_digits',check_affiliations_list[index:country_pos]))
                                break_id = 'droping_digits'
                                if verbose:
                                    print('Break identification:', break_id, '\n')
                                break
                            else:
                                if not add_affiliation_flag: 
                                    affiliations_list.append(affiliation)
                                    add_affiliation_flag = True
                                    break_id = 'droping_digits aborted by user_keeping_words'
                                    if verbose:
                                        print('Break identification:', break_id, '\n')
                    if found_item_flags.droping_town:
                        if len(check_affiliations_list[index:country_pos])<=2:   
                            affiliations_drop.append(('droping_town', check_affiliations_list[index:country_pos]))
                            break_id = 'droping_town'
                            if verbose:
                                print('Break identification:', break_id, '\n')    
                            break
                        else:
                            affiliations_drop.append(('droping_town', affiliation))
                            break_id = 'droping_town aborted by index of town in affiliations list'
                            if verbose:
                                print('Break identification:', break_id, '\n')
                    else:
                        if found_item_flags.droping_suffix:
                            if found_item_flags.gen_keeping_words or found_item_flags.user_keeping_words:
                                if not add_affiliation_flag: 
                                    affiliations_list.append(affiliation)
                                    add_affiliation_flag = True
                                if found_item_flags.gen_keeping_words: break_id = 'droping_suffix aborted by gen_keeping_words'
                                if found_item_flags.user_keeping_words: break_id = 'droping_suffix aborted by user_keeping_words'
                                if verbose:
                                    print('Break identification:', break_id, '\n')    
                            else:
                                affiliations_drop.append(('droping_suffix',check_affiliations_list[index:country_pos]))
                                break_id = 'droping_suffix'
                                if verbose:
                                    print('Break identification:', break_id, '\n')
                                break   
                        if found_item_flags.droping_words:
                            # Keeping affiliation when a keeping word is found only if no droping digit is found
                            # this keeps "department bldg civil" which is wanted even if "bldg" is a droping word 
                            # unfortunatly, this keeps unwanted "campus university", "ciudad university"... 
                            if any(keeping_words_flags) and not found_item_flags.droping_digits:  
                                if not add_affiliation_flag: 
                                    affiliations_list.append(affiliation)
                                    add_affiliation_flag = True
                                if found_item_flags.user_keeping_words: break_id = 'droping_word aborted by user_keeping_words'
                                if found_item_flags.basic_keeping_words: break_id = 'droping_word aborted by basic_keeping_words'
                                if found_item_flags.gen_keeping_words: break_id = 'droping_word aborted by gen_keeping_words'
                                if verbose:
                                    print('Break identification:', break_id, '\n')
                            else:
                                # Droping affiliation from affiliations_list if already added because of a former drop abort
                                if add_affiliation_flag:
                                    affiliations_list = affiliations_list[:-1]
                                    add_affiliation_flag = False
                                affiliations_drop.append(('droping_words', check_affiliations_list[index:country_pos]))
                                break_id = 'droping_words'
                                if verbose:
                                    print('Break identification:', break_id, '\n')
                                break             
    else:
        affiliations_list = raw_affiliations_list
        affiliations_drop = []

    # Removing spaces from the kept affiliations 
    affiliations_list = [x.strip() for x in affiliations_list]
    if verbose:
        print('affiliations_list stripped:', affiliations_list, "\n")
        
    # Removing country and country alias from the kept affiliations 
    affiliations_list = [x for x in affiliations_list if x != country and x not in bp_gg.ALIAS_UK]        
    if verbose:
        print('affiliations_list without country aliases:', affiliations_list, "\n")
    
    return country, affiliations_list, affiliations_drop


def _build_words_set(raw_aff, verbose=False):
    """Builds sets of words from a raw affiliation after standardization of words and symbols, 
    removing special symbols, adding missing spaces and droping small words.

    Args:
        raw_aff (str): The raw affiliation used to build the sets of words.
        verbose (bool): If true, variables are printed for code control (default: False).
    Returns:
        (tuple): Tuple of to sets of words; The first set is the canonical set of words \
        issuing from the string 'raw_aff'; The second set is an added set if some specific \
        accronyms are present in the first set of words.
    """   
    # Setting substitution templates for searching small words or acronyms
    small_words_template = Template(r'[\s(]$word[\s)]' # For instence capturing 'of' in 'technical university of denmark'                                                              
                                    + '|'
                                    + r'[\s]$word$$' # For instence capturing 'd' in 'institut d ingenierie'
                                    + '|'
                                    + r'^$word\b') # For instence capturing 'the' in 'the denmark university'

    accronymes_template = Template(r'[\s(]$word[\s)]' # For instence capturing 'umr' in 'umr dddd' or 'umr dd'                                                             
                                  + '|'
                                  + r'[\s]$word$$'      
                                  + '|'
                                  + r'^$word\b')

    # Removing accents and spaces at ends
    raw_aff_mod = remove_special_symbol(raw_aff, only_ascii=False, strip=True)

    # Uniformizing words
    std_raw_aff = raw_aff_mod
    std_raw_aff_add = ""
    for word_to_substitute, re_pattern in bp_sg.DIC_WORD_RE_PATTERN.items():
        std_raw_aff = re.sub(re_pattern, word_to_substitute + ' ', std_raw_aff)
    std_raw_aff = re.sub(r'\s+', ' ', std_raw_aff)
    std_raw_aff = re.sub(r'\s,', ',', std_raw_aff)
    std_raw_aff = std_raw_aff.lower()

    # Uniformizing dashes
    std_raw_aff = std_raw_aff.translate(bp_gg.DASHES_CHANGE)

    # Uniformizing apostrophes
    std_raw_aff = std_raw_aff.translate(bp_gg.APOSTROPHE_CHANGE)

    # Uniformizing symbols
    std_raw_aff = std_raw_aff.translate(bp_gg.SYMB_CHANGE)

    # Droping particular symbols
    std_raw_aff = std_raw_aff.translate(bp_gg.SYMB_DROP)
    if verbose:
        print('       std_raw_aff:', std_raw_aff)

    # Building the corresponding set of words to std_raw_aff
    raw_aff_words_set = set(std_raw_aff.strip().split(' '))

    # Managing missing space in raw affiliations related 
    # to particuliar institutions cases such as UMR or U followed by digits
    for accron in bp_sg.MISSING_SPACE_ACRONYMS:
        re_accron = re.compile(accronymes_template.substitute({"word":accron}))
        if re.search(re_accron,std_raw_aff.lower()) and len(raw_aff_words_set)==2:
            std_raw_aff_add = "".join(std_raw_aff.split(" "))

    # Droping small words
    for word_to_drop in bp_sg.SMALL_WORDS_DROP:
        re_drop_words = re.compile(small_words_template.substitute({"word":word_to_drop}))
        if re.search(re_drop_words,std_raw_aff.lower()):
            raw_aff_words_set = raw_aff_words_set - {word_to_drop}

    # Updating raw_aff_words_set_list using std_raw_aff_add
    raw_aff_words_set_add = {}
    if std_raw_aff_add:
        raw_aff_words_set_add = set(std_raw_aff_add.split(' '))

    return raw_aff_words_set, raw_aff_words_set_add


def _build_words_sets_list(raw_aff_list, verbose=False): 
    """Builds a list of words sets from a list of raw affiliations.

    Args: 
        raw_aff_list (list): The list of raw affiliations as strings.
        verbose (bool): If true, variables are printed for code control (default: False).
    Returns:
        (list): List of words sets.
    """
    raw_aff_words_sets_list = []
    for idx, raw_aff in enumerate(raw_aff_list):
        if raw_aff and raw_aff!=' ':                
            # Building the set of words for raw affiliation
            raw_aff_words_set, raw_aff_words_set_add = _build_words_set(raw_aff, verbose)

            # Upadating the list of words sets with the set raw_aff_words_set 
            raw_aff_words_sets_list.append(raw_aff_words_set)

            # Upadating the list of words sets using the set raw_aff_words_set_add
            if raw_aff_words_set_add: raw_aff_words_sets_list.append(raw_aff_words_set_add)

    return raw_aff_words_sets_list


def build_norm_raw_affiliations_dict(country_affiliations_file_path=None, verbose=False):
    """Builds a dict keyyed by country and the value per country is a dict keyyed 
    by normalized affiliation and valued by a list of sets of words representing 
    the raw affiliations corresponding to the normalized affiliation.

    Args:
        country_affiliations_file_path (path): Full path to the file of normalized affiliations \
        with they possible corresponding raw affiliation built by the user"; if None, it is set \
        using the 'COUNTRY_AFFILIATIONS_FILE' and 'REP_UTILS' globals.
        verbose (bool): If true, variables are printed for code control (default: False).
    Returns:
        (dict): The built dict.
    """
    # Setting the path for the 'Country_affilialions.xlsx' file
    if not country_affiliations_file_path:
        country_affiliations_file_path = Path(bp.__file__).parent / Path(bp_gg.REP_UTILS) / Path(bp_sg.COUNTRY_AFFILIATIONS_FILE)
   
    # Reading the 'Country_affilialions.xlsx' file in the dataframe dic    
    wb = openpyxl.load_workbook(country_affiliations_file_path)
    country_aff_df = pd.read_excel(country_affiliations_file_path, 
                                   sheet_name = wb.sheetnames)

    norm_raw_aff_dict = {}
    for country_aff_df_item in country_aff_df.items():
        country = country_aff_df_item[0]
        norm_raw_aff_dict[country] = {}
        norm_raw_aff_df = country_aff_df_item[1]    
        norm_raw_aff_nb = len(norm_raw_aff_df["Norm affiliations"])

        if verbose:
            print('Country:', country)
            print('Number of norm affiliations:', norm_raw_aff_nb)
            print('\nList of norm affiliations:', norm_raw_aff_df["Norm affiliations"], "\n")

        for num, norm_aff in enumerate(norm_raw_aff_df["Norm affiliations"]):
            norm_aff = norm_aff.strip()
            raw_aff_list = [item for item in list(norm_raw_aff_df.loc[num])[1:] if not(pd.isnull(item))==True]

            if verbose:
                print(f"\n\n{str(num)}- Norm affiliation: {norm_aff}")
                print('   Raw affiliations list:', raw_aff_list, "\n")

            norm_raw_aff_dict[country][norm_aff] = _build_words_sets_list(raw_aff_list, verbose)

            if verbose:
                print(f"   norm_raw_aff_dict[{country}][{norm_aff}]: {norm_raw_aff_dict[country][norm_aff]}\n")
            
    return norm_raw_aff_dict


def read_inst_types(inst_types_file_path=None, inst_types_usecols=None):
    """Builds a dict keyyed by normalized affiliations types and the value per type 
    is the order level of the type.
   
    Args:
        inst_types_file_path (path): The full path to the file of ordered institutions types; \
        if None, it is set using the 'INST_TYPES_FILE' and 'REP_UTILS' globals.
        inst_types_usecols (list of str): The list of columns names for order levels \
        and abbreviations of affiliation types in the file "Inst_types.xlsx"; f None, \
        it is set using the 'INST_TYPES_USECOLS' global.
    Returns:
        (dict): The built dict.
    """
    # Setting the full path for the file of ordered institutions types
    if not inst_types_file_path:
        inst_types_file = bp_sg.INST_TYPES_FILE
        inst_types_file_path = Path(bp.__file__).parent / Path(bp_gg.REP_UTILS) / Path(inst_types_file)
        
    if not inst_types_usecols:
        inst_types_usecols = bp_sg.INST_TYPES_USECOLS

    # Reading the file in a dataframe
    inst_types_df = pd.read_excel(inst_types_file_path, usecols=bp_sg.INST_TYPES_USECOLS)

    levels = [x for x in inst_types_df['Level']]
    abbreviations = [x for x in inst_types_df['Abbreviation']]
    aff_type_dict = dict(zip(abbreviations, levels))

    return aff_type_dict


def read_towns_per_country(country_towns_file=None, country_towns_folder_path=None):
    """Builds dict keyyed by countries and valued by a list of towns of the each country.

    It uses the functions `rationalize_town_names`and `remove_special_symbol`
    imported from the `BiblioParsing.BiblioParsingUtils` module.
    
    Args:
        country_towns_file (str): File name of the list of towns per country.
        country_towns_folder_path (path): The full path to the folder \
        of the 'country_towns_file' file.
        
    Returns:
        (dict): The built dict.
    """
    # Setting the path of the file of towns par country
    if not country_towns_folder_path:
        country_towns_folder_path = Path(bp.__file__).parent / Path(bp_gg.REP_UTILS)
    if not country_towns_file:
        file_path = country_towns_folder_path / Path(bp_sg.COUNTRY_TOWNS_FILE)
    else:
        file_path = country_towns_folder_path / Path(country_towns_file)

    # Reading the file of towns per country in a dict of dataframes 
    wb = openpyxl.load_workbook(file_path)
    df_dict = pd.read_excel(file_path, 
                            sheet_name=wb.sheetnames)
    
    towns_dict = {x[0]:x[1]['Town name'].tolist() for x in df_dict.items()}
    for country in towns_dict.keys():        
        list_towns = []
        for town in towns_dict[country]:
            town = town.lower()
            town = rationalize_town_names(town, dic_town_symbols=bp_sg.DIC_TOWN_SYMBOLS,
                                          dic_town_words=bp_sg.DIC_TOWN_WORDS)
            town = remove_special_symbol(town, only_ascii=False, strip=False)
            town = town.strip()
            list_towns.append(town)
        towns_dict[country] = list_towns    
    return towns_dict


def _check_norm_raw_aff_dict(norm_raw_aff_dict, aff_type_dict, user_country_affiliations_file_path):
    wrong_affil_types_dict = {}
    aff_types_set = set(aff_type_dict.keys())
    for country, country_dict in norm_raw_aff_dict.items():
        norm_affil_types_set = set([norm_affil.split(' ')[-1] for norm_affil in country_dict.keys()])
        norm_affil_types_in = aff_types_set.intersection(norm_affil_types_set)
        norm_affil_types_out = norm_affil_types_set - norm_affil_types_in
        if norm_affil_types_out:
            wrong_affil_types_dict[country] = list(norm_affil_types_out)
    return wrong_affil_types_dict


def build_norm_raw_institutions(addresses_df, inst_types_file_path=None, country_affiliations_file_path=None,
                                country_towns_file=None, country_towns_folder_path=None,
                                verbose=False, progress_param=None):    
    """Parses the addresses of each publication of the corpus to retrieve the country, 
    the normalized institutions and the institutions not yet normalized for each address.

    Args:
        addresses_df (dataframe): the data of the addresses resulting from the parsing of \
        the corpus after concatenation and deduplication of partial parsings.
        inst_types_file_path (path): The full path to the data of institutions-types used to normalize \
        the affiliations, optional (default=None).
        country_affiliations_file_path (path): The full path to the data per country of raw affiliations \
        per normalized one, optional (default=None).
        country_towns_file (str): The name of the file of the data of towns per country, optional (default=None).
        country_towns_folder_path (path): The full path to the folder where the 'country_towns_file' file \
        is available, optional (default=None).
        verbose (bool): If set to 'True' allows prints for code control (default: False).
        progress_param (tup): (Function for updating ProgressBar tkinter widget status, \
        The initial progress status (int), The final progress status (int)) \
        (optional, default = None)
    Returns:
        (tuple): (countries data per address (dataframe), normalized affiliations per address (dataframe), \
        raw institutions per address (dataframe), A dict of wrong type of normalized affiliation \
        for correction by the user).
    """
    # Setting useful column names
    cols_lists_dic, cols_dic = _set_norm_affiliations_cols()    
    cols_lists_keys = ['country_cols_list', 'inst_cols_list']
    country_cols_list, inst_cols_list = [cols_lists_dic[key] for key in cols_lists_keys]    
    cols_keys = ['pub_id_col', 'address_col', 'country_col', 'institution_col']
    pub_id_col, address_col, country_col, institution_col = [cols_dic[key] for key in cols_keys]
    
    # Setting useful cols lists
    norm_inst_cols_list = inst_cols_list
    raw_inst_cols_list = inst_cols_list + [address_col]
    
    # Setting named tuples
    country= namedtuple('country', country_cols_list)
    norm_institution = namedtuple('norm_institution', norm_inst_cols_list)
    raw_institution = namedtuple('raw_institution', raw_inst_cols_list)
    
    # Getting useful dicts for affiliation normalization
    aff_type_dict = read_inst_types(inst_types_file_path=inst_types_file_path,
                                    inst_types_usecols=None)
    norm_raw_aff_dict = build_norm_raw_affiliations_dict(country_affiliations_file_path=country_affiliations_file_path)
    towns_dict = read_towns_per_country(country_towns_file=country_towns_file,
                                        country_towns_folder_path=country_towns_folder_path)
    wrong_affil_types_dict = _check_norm_raw_aff_dict(norm_raw_aff_dict, aff_type_dict,
                                                      country_affiliations_file_path)

    if not wrong_affil_types_dict:
        step_nb = len(addresses_df)
        step = 0
        if progress_param:
            progress_callback, init_progress, final_progress = progress_param
            progress_step = (final_progress-init_progress) / step_nb
            progress_status = init_progress
            progress_callback(progress_status)

        countries_list = []
        norm_institutions_list = []
        raw_institutions_list = []
        for pub_id, address_dg in addresses_df.groupby(pub_id_col):
            if verbose:
                print("\n\nPub_id:", pub_id)
                print("\naddress_dg:\n", address_dg)     
            for idx, raw_address in enumerate(address_dg[address_col].tolist()):
                std_address = standardize_address(raw_address)
                address_country = ""
                address_norm_affiliation_list = []
                address_raw_affiliation_list = []            
                try:
                    aff_list_tup = _build_address_affiliations_lists(std_address, norm_raw_aff_dict,
                                                                     aff_type_dict, towns_dict,
                                                                     drop_status=True, verbose=False)
                    address_country, address_norm_affiliation_list, address_raw_affiliation_list = aff_list_tup
                except KeyError:
                    print("\n\nError Pub_id / idx:", pub_id," / ", idx)
                    print("\naddress_dg:\n", address_dg[address_col].tolist()[idx])
                    pass
                address_norm_affiliations = bp_sg.EMPTY
                address_raw_affiliations = bp_sg.EMPTY
                if address_norm_affiliation_list: 
                    address_norm_affiliations = "; ".join(address_norm_affiliation_list)
                if address_raw_affiliation_list: 
                    address_raw_affiliations = "; ".join(address_raw_affiliation_list)
                if address_country:
                    countries_list.append(country(pub_id, idx, address_country))
                norm_institutions_list.append(norm_institution(pub_id, idx, address_norm_affiliations))
                raw_institutions_list.append(raw_institution(pub_id, idx, address_raw_affiliations, std_address))
                step += 1

                if verbose:
                    print('\nIdx address:                       ', idx)
                    print('Country:                           ', address_country)
                    print('address_norm_affiliation_list:     ', address_norm_affiliations)
                    print('address_unknown_affiliations_list: ', address_raw_affiliations)
                    print(f"        Number of addresses analyzed: {step} / {step_nb}")
                else:
                    print(f"        Number of addresses analyzed: {step} / {step_nb}", end="\r")

                if progress_param:
                    progress_status += progress_step
                    progress_callback(progress_status)

        # Building a clean countries dataframe and accordingly updating the parsing success rate dict
        country_df, _ = build_item_df_from_tup(countries_list, country_cols_list,
                                               country_col, pub_id_col)

        # Building a clean institutions dataframe and accordingly updating the parsing success rate dict
        norm_institution_df, _ = build_item_df_from_tup(norm_institutions_list, norm_inst_cols_list,
                                                        institution_col, pub_id_col)

        # Building a clean institutions dataframe and accordingly updating the parsing success rate dict
        raw_institution_df, _ = build_item_df_from_tup(raw_institutions_list, raw_inst_cols_list,
                                                       institution_col, pub_id_col)
    else:
        # Returning empty dataframes
        country_df = pd.DataFrame()
        norm_institution_df= pd.DataFrame()
        raw_institution_df = pd.DataFrame()

    if progress_param:
        progress_callback, _, final_progress = progress_param
        progress_callback(final_progress)

    return country_df, norm_institution_df, raw_institution_df, wrong_affil_types_dict
